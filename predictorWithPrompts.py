# -*- coding: utf-8 -*-
"""
Created on Mon Feb 19 14:25:18 2024

@author: Mohammed
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 01:47:37 2024

@author: Mohammed
"""

import sys
from time import sleep

using_colab = False
import numpy as np
import torch
import matplotlib.pyplot as plt
import cv2
from openpyxl import Workbook
import statistics
import os
import random
from IPython import get_ipython
from matplotlib.backend_bases import MouseButton
import time
from PIL import Image
from segment_anything import sam_model_registry, SamPredictor
import matplotlib

plt.rcParams['keymap.grid'].remove('g')
plt.rcParams['keymap.home'].remove('r')

MEDIUM_STAR_SIZE = 50 
MEDIUM_GREEN_RED_DOT_SIZE = 5
SMALL_STAR_SIZE = 10
SMALL_GREEN_RED_DOT_SIZE = 2

MEDIUM_DOT_SIZE_MODE = False
SMALL_DOT_SIZE_MODE = True
dot_size_toggle = SMALL_DOT_SIZE_MODE # small dot size by default
GREEN_COLOR = '#00f700'
RED_COLOR = '#ff1919'


def show_mask(mask, ax, random_color=False):
    if random_color:
        color = np.concatenate([np.random.random(3), np.array([0.6])], axis=0)
    else:
        color = np.array([30 / 255, 144 / 255, 255 / 255, 0.6])
    h, w = mask.shape[-2:]
    mask_image = mask.reshape(h, w, 1) * color.reshape(1, 1, -1)
    ax.imshow(mask_image)


def show_points(coords, labels, ax, marker_size=50):
    pos_points = coords[labels == 1]
    neg_points = coords[labels == 0]
    
    if dot_size_toggle == MEDIUM_DOT_SIZE_MODE:
        ax.scatter(pos_points[:, 0], pos_points[:, 1], color=GREEN_COLOR, marker='*', s=marker_size, edgecolor='white',
                linewidth=1.25)
        ax.scatter(neg_points[:, 0], neg_points[:, 1], color=RED_COLOR, marker='*', s=marker_size, edgecolor='white',
               linewidth=1.25)
    else:
        ax.scatter(pos_points[:, 0], pos_points[:, 1], color=GREEN_COLOR, marker='*', s=marker_size, edgecolor='white',
                linewidth=0.5)
        ax.scatter(neg_points[:, 0], neg_points[:, 1], color=RED_COLOR, marker='*', s=marker_size, edgecolor='white',
               linewidth=0.5)


def closetn(node, nodes):
    nodes = np.asarray(nodes)
    deltas = nodes - node
    dist_2 = np.einsum('ij,ij->i', deltas, deltas)
    return np.argmin(dist_2)


sys.path.append("..")

try:
    matplotlib.use('Qt5Agg')
except:
    matplotlib.use('TkAgg')

sam_checkpoint = 'sam_vit_h_4b8939.pth'
model_type = "vit_h"

device = "cuda" if torch.cuda.is_available() else 'cpu'

sam = sam_model_registry[model_type](checkpoint=sam_checkpoint)
sam.to(device=device)

predictor = SamPredictor(sam)

names = np.load("samples.npy", allow_pickle=True)
labels = np.load("labels.npy", allow_pickle=True)

dataset_name = input("What dataset are you running?")
name = dataset_name + "_Timed"
if not os.path.exists(name):
    os.makedirs(name)


# %%

#first = input("Do you want to load previous work? -y -n\n")
#while first != 'n' and first != 'y':
#    first = input("Chose y or n, Do you want to load previous work? -y -n\n")
#if first == 'n':
#    wb = Workbook()
#    ws = wb.active
#    ws['A1'] = 'Scores'
#    ws['B1'] = 'Times'
#    ws['C1'] = '# red dots of best '
#    ws['D1'] = 'SD of green of best '
#    ws['E1'] = 'SD of red of best'
#    ws['F1'] = 'best score'
    # for i in range(10):
    #     ws[i+'1']='# green dots of '+str(i)
    #     ws[chr(72+i*5)+'1']='# red dots of '+str(i)
    #     ws[chr(73+i*5)+'1']='SD of green of '+str(i)
    #     ws[chr(74+i*5)+'1']='SD of red of '+str(i)
    #     ws[chr(75+i*5)+'1']='score of '+str(i)
 #   serv=np.array([])
  #  for i in range(9):
   #     coun = 1
    #    for col in ws.iter_cols(min_row=1, max_row=1, max_col=12 + i * 5, min_col=7 + i * 5):
     #       if coun == 1:
      #          ws[col[0].coordinate] = '# green dots of ' + str(i + 2)
       #     elif coun == 2:
        #        ws[col[0].coordinate] = '# red dots of ' + str(i + 2)
         #   elif coun == 3:
#                ws[col[0].coordinate] = 'SD of X of ' + str(i + 2)
 #           elif coun == 4:
  #              ws[col[0].coordinate] = 'SD of Y of ' + str(i + 2)
   #         elif coun == 5:
    #            ws[col[0].coordinate] = 'score of ' + str(i + 2)
#            coun += 1
 #   name = input("Type your name:\n")
#
  #  if not os.path.exists(name):
 #       os.makedirs(name)
  #      os.makedirs(os.path.join(name, "masks"))
   #     os.makedirs(os.path.join(name, "points"))
    #    os.makedirs(os.path.join(name, "sorts"))
     #   os.makedirs(os.path.join(name, "eachround"))
      #  os.makedirs(os.path.join(name, "scores"))

 #   c = 0
  #  tim = 0
  #  t = time.time()
#else:
#    from openpyxl import load_workbook
    
#    name = input("what is your name?\n")
#    wb = load_workbook(os.path.join(name, name + ".xlsx"))
#    ws = wb.active
   # c = len(os.listdir(os.path.join(name, "masks")))
    #f = open(os.path.join(name, "time.txt"), 'r')
#    serv=np.load(os.path.join(name,"servey.npy")) if os.path.exists(os.path.join(name,"servey.npy")) else np.array([])
#    tim = f.readline()
#    t = time.time()
#    f.close()

#### change that later

c=0

f = False
## start looping through samples:
for student in np.arange(1,2): 
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Sample Number'
    ws['B1'] = 'Scores'
    ws['C1'] = 'Times'
    c=0
    num_files = 0
    dir_path = dataset_name + "/st" + str(student) + "/scores/"
    # Iterate directory
    for path in os.listdir(dir_path):
        # check if current path is a file
        if os.path.isfile(os.path.join(dir_path, path)):
            num_files += 1
    while c < 5 and not f:
        score_points = np.load("Cat/st" + str(student) + "/scores/" + str(c) + "score.npy", allow_pickle=True)
        if score_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        green_points = np.load("Cat/st" + str(student) + "/points/" + str(c) + "_green.npy", allow_pickle=True)
        if green_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        green_points = green_points[np.argmax(score_points)]
        red_points = np.load("Cat/st" + str(student) + "/points/" + str(c) + "_red.npy", allow_pickle=True)
        if red_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        red_points = red_points[np.argmax(score_points)]
        msk = []  # masks for each samples

        gp = []  # green points
        rp = []  # red points
        image = names[c]  # samples c
        if len(image.shape) == 2:
            image = cv2.cvtColor((np.array(((image + 1) / 2) * 255, dtype='uint8')), cv2.COLOR_GRAY2RGB)
        label = labels[c]  # GT for sample c
        rmv = False
        mask = 0
        # image=np.array(((image+1)/2)*255,dtype='uint8')
         
        predictor.set_image(image)
        inc = ""
        co = 0
        bs = 0
        score = []
        round=[0,0]
        stdx = []
        stdy = []
        ng = []
        nr = []
        green = []
        red = []
        greenx = []
        times = []
        
        redx = []
        greeny = []
        redy = []
        
        # label=plt.imread('C:/Users/Mohammed/Downloads/labels/'+labels[c])i9i
        label = label == 1

        # matplotlib.use('TkAgg')

        while inc != "y":
            s = 0  # this is for the score
            count = 1  # to count the score max
            lessfive = 0
            current_color = 'green'
            dot_size_toggle = SMALL_DOT_SIZE_MODE # default will be small dot, not medium
            current_star_size = SMALL_STAR_SIZE
            current_green_red_dot_size = SMALL_GREEN_RED_DOT_SIZE
            # get_ipython().run_line_magic('matplotlib', 'qt')



            def addPoints():
                x,y = 0,0
                for g in green_points:
                    x = g[0]
                    y = g[1]
                    green.append((x, y))
                    greenx.append(x)

                    greeny.append(y)
                
                

                for r in red_points:
                    x = r[0]
                    y = r[1]
                    red.append((x, y))
                    redx.append(x)

                    redy.append(y)
            
                if green and red:
                    global s
                    #print("green:", green)
                    #print("red:", red)

                    input_point = np.concatenate((green, red))
                    input_label = np.concatenate(([1] * len(green), [0] * len(red)))

                    time_1 = time.time()
                    
                    
                    # Get the time in seconds 
                    # since the epoch 
                    
                    masks, scores, logits = predictor.predict(
                        point_coords=input_point,
                        point_labels=input_label,
                        multimask_output=True,
                    )
                    time_sec = time.time() - time_1
                    times.append(time_sec)
                    
                    # Print the time  
                # print("Time in seconds since the epoch:", time_sec) 
                    mask = masks[0]

                    # get_ipython().run_line_magic('matplotlib', 'inline')
                    
                    intersection = (mask & label).sum()
                    union = (mask | label).sum()
                    if intersection == 0:
                        s = 0
                    else:
                        s = intersection / union
                    # ws[chr(68)+str(c+2)]=str(bs) # start at cell D(c)
                    #show_points(input_point, input_label, ax[2], marker_size = current_star_size)
                    msg = ""

                    if len(score[round[0]:]) == 0:
                        maxx = 0
                    else:
                        maxx = max(score[round[0]:])
                        print("maxx",maxx)
                    score.append(s)
                    gp.append(np.multiply(green, 1))

                    rp.append(np.multiply(red, 1))
                    ng.append(len(greenx))
                    nr.append(len(redx))
                    grx = np.concatenate([greenx, redx])
                    gry = np.concatenate([greeny, redy])

                    stdx.append(statistics.pstdev(grx.astype(int).tolist()))
                    stdy.append(statistics.pstdev(gry.astype(int).tolist()))
                
                    plt.title(f"Score: {(intersection / union):.3f}" + msg, fontsize=13)
                    ## saving masks, scores, points and other stats: 
                    msk.append(np.multiply(mask, 5))
                    #print("less than best score", lessfive)
                    #print("scores:", score[0])
                    #print("times:", times[0] )


                    



        
                    
                    

            # Create a figure and display the image

            #ax[0].imshow(image)
            #ax[1].imshow(label)
            # Connect mouse click and keyboard key events
        
            # fig.canvas.start_event_loop(timeout=-5)
            # fig.canvas.start_event_loop(timeout=-5)
            # Display the plot

            #fig.show()  # this call does not block on my system

            #fig.canvas.start_event_loop()  # block here until window closed
            addPoints()
            # After closing the image window, you can access the green and red pixel coordinate lists

            # To select the truck, choose a point on it. Points are input to the model in (x,y) format and come with labels 1 (foreground point) or 0 (background point). Multiple points can be input; here we use only one. The chosen point will be shown as a star on the image.
            # print("Hereeeeeeeee")

            # ws['B'+str(c+2)]=str(len(green)) 
            # ws['C'+str(c+2)]=str(len(red))
            # ws['D'+str(c+2)]=str()
            # input_point=np.concatenate((green,red))
            # input_label=np.concatenate(([1]*len(green),[0]*len(red)))

            # masks, scores, logits = predictor.predict(
            #     point_coords=input_point,
            #     point_labels=input_label,
            #     multimask_output=True,
            # )

            # sleep(1)
            # if np.max(score)<0.8:
            #     print("your score should be more than 0.8, try again")
            #     inc=""
            #     co+=1
            #     if co>=2:
            #         inc=input("you tried more than 10 times\nYou can continue and save the best score ("+str(max(score))+")\nif you want to continue press y")
            # else:
            inc = "y"
            #print(inc)

        indx = np.argsort(-np.array(score))
        sscore = np.array(score)[indx]
        times_2 = np.array(times)[indx]
        snr = np.array(nr)[indx]
        sstdx = np.array(stdx)[indx]
        sstdy = np.array(stdy)[indx]
        ws['A' + str(c + 2)] = str(c)
        ws['B' + str(c + 2)] = sscore[0]  # samples name on excel
        ws['C' + str(c + 2)] = times_2[0]
        #for i in range(len(score)):
        #    coun = 1
        #    for col in ws.iter_cols(min_row=c + 2, max_row=c + 2, max_col=6 + i * 5, min_col=2 + i * 5):
        #        if coun == 1:
        #            ws[col[0].coordinate] = sscore[i]
        #        elif coun == 2:
        #            ws[col[0].coordinate] = times_2[i]
        #        elif coun == 3:
        #            ws[col[0].coordinate] = sstdx[i]
        #        elif coun == 4:
        #            ws[col[0].coordinate] = sstdy[i]
        #        elif coun == 5:
        #            ws[col[0].coordinate] = sscore[i]
        #        coun += 1
        #np.save(os.path.join(name, "points", str(c) + "_green"), np.array(gp, dtype=object))
        #np.save(os.path.join(name, "points", str(c) + "_red"), np.array(rp, dtype=object))
        #np.save(os.path.join(name, "masks", str(c) + "_mask"), np.array(msk))
        #np.save(os.path.join(name, "sorts", str(c) + "_sort"), indx)
        #np.save(os.path.join(name, "scores", str(c) + "score"), score)
        #np.save(os.path.join(name, "times", str(c) + "time"), time)
        #np.save(os.path.join(name,"eachround",str(c)+"_"),round)

        c += 1
        #contin = input("do u want to continue? press y if you want to continue or anyting otherwise ")
        #if not contin == 'y':
            #wb.save(os.path.join(name, name + '.xlsx'))
            #f = True
            # file = open(os.path.join(name, "time.txt"), 'w')
            # file.write(str(float(tim) + (time.time() - t)))
            # file.close()
        print("Sample:", c)
    wb.save(os.path.join(name, 'st' + str(student) + name + '.xlsx'))


def addPoints():
                x,y = 0,0
                
                seen = set(green)
                for g in green_points:
                    x = g[0] / 2**int(compressions)
                    y = g[1] / 2**int(compressions)
                    
                    if ((x,y)) not in green:
                        seen.add((x,y))
                        green.append((x, y))
                        greenx.append(x)
                        greeny.append(y)
                
                seen = set(red)
                for r in red_points:
                    x = r[0] / 2**int(compressions)
                    y = r[1] / 2**int(compressions)
                    if ((x,y)) not in seen:
                        seen.add((x,y))
                        red.append((x, y))
                        redx.append(x)
                        redy.append(y)
            
                if green and red:
                    global s
                    #print("green:", green)
                    #print("red:", red)

                    input_point = np.concatenate((green, red))
                    input_label = np.concatenate(([1] * len(green), [0] * len(red)))

                    time_part2i = time.time()
                    
                    
                    # Get the time in seconds 
                    # since the epoch 
                    
                    masks, scores, logits = predictor.predict(
                        point_coords=input_point,
                        point_labels=input_label,
                        multimask_output=True,
                    )
                    time_part2 = time.time() - time_part2i
                    times.append(time_part1 + time_part2)
                    
                    # Print the time  
                # print("Time in seconds since the epoch:", time_sec) 
                    mask = masks[0]
                    maxunPool = torch.nn.Upsample(size=imshape, mode='nearest')
                    imageTensor = torch.from_numpy(mask)
                    imageTensor = imageTensor.float().unsqueeze(dim=0)
                    imageTensor = imageTensor.float().unsqueeze(dim=0)
                    imageTensorDone = maxunPool(imageTensor)
                    no_batch = imageTensorDone.long().squeeze(dim=0)
                    imTens = no_batch.long().squeeze(dim=0)
                    imageUp = imTens.numpy()
                    mask = imageUp.astype(int)

                    # get_ipython().run_line_magic('matplotlib', 'inline')
                    
                    intersection = (mask & label).sum()
                    union = (mask | label).sum()
                    if intersection == 0:
                        s = 0
                    else:
                        s = intersection / union
                    # ws[chr(68)+str(c+2)]=str(bs) # start at cell D(c)
                    #show_points(input_point, input_label, ax[2], marker_size = current_star_size)
                    msg = ""

                    if len(score[round[0]:]) == 0:
                        maxx = 0
                    else:
                        maxx = max(score[round[0]:])
                        print("maxx",maxx)
                    score.append(s)
                    gp.append(np.multiply(green, 1))

                    rp.append(np.multiply(red, 1))
                    ng.append(len(greenx))
                    nr.append(len(redx))
                    grx = np.concatenate([greenx, redx])
                    gry = np.concatenate([greeny, redy])

                    stdx.append(statistics.pstdev(grx.astype(int).tolist()))
                    stdy.append(statistics.pstdev(gry.astype(int).tolist()))
                
                    plt.title(f"Score: {(intersection / union):.3f}" + msg, fontsize=13)
                    ## saving masks, scores, points and other stats: 
                    msk.append(np.multiply(mask, 5))
                    #print("less than best score", lessfive)
                    #print("scores:", score[0])
                    #print("times:", times[0] )
