# -*- coding: utf-8 -*-
"""
Created on Mon Feb 19 14:25:18 2024

@author: Mohammed
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 01:47:37 2024

@author: Mohammed
"""

import sys
from time import sleep

using_colab = False
import numpy as np
import torch
import matplotlib.pyplot as plt
import cv2
from openpyxl import Workbook
import statistics
import os
import random
from IPython import get_ipython
from matplotlib.backend_bases import MouseButton
import time
from PIL import Image
from segment_anything import sam_model_registry, SamPredictor
import matplotlib

def show_mask(mask, ax, random_color=False):
    if random_color:
        color = np.concatenate([np.random.random(3), np.array([0.6])], axis=0)
    else:
        color = np.array([30 / 255, 144 / 255, 255 / 255, 0.6])
    h, w = mask.shape[-2:]
    mask_image = mask.reshape(h, w, 1) * color.reshape(1, 1, -1)
    ax.imshow(mask_image)


sys.path.append("..")

sam_checkpoint = 'sam_vit_h_4b8939.pth'
model_type = "vit_h"

device = "cuda" if torch.cuda.is_available() else 'cpu'

sam = sam_model_registry[model_type](checkpoint=sam_checkpoint)
sam.to(device=device)

predictor = SamPredictor(sam)

########################## Set User Parameters ##########################
dataset_name = "Dolphin below" # "Cat", "Bus", "Dolphin below"
max_samples = 5
blur_ratio = 7
#########################################################################

names = np.load(dataset_name+"/samples.npy", allow_pickle=True)
labels = np.load(dataset_name+"/labels.npy", allow_pickle=True)
down_name = dataset_name + "_Timed_DownSampled"
if not os.path.exists(down_name):
    os.makedirs(down_name)

#### change that later

c=0
f = False
## start looping through samples:
for student in np.arange(1,2):
    print("Student: ", student) 
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Sample Number'
    ws['B1'] = 'Scores'
    ws['C1'] = 'Time Set'
    ws['D1'] = 'Time Predict'
    
    # Count the number of samples
    num_files = 0
    dir_path = dataset_name + "/st" + str(student) + "/scores/"
    # Iterate directory
    for path in os.listdir(dir_path):
        # check if current path is a file
        if os.path.isfile(os.path.join(dir_path, path)):
            num_files += 1       
    
    # Test for Different Scale ratio            
    c=0
    while c < num_files and c < max_samples and not f:
        print("Sample:", c)
        score_points = np.load(dataset_name + "/st" + str(student) + "/scores/" + str(c) + "score.npy", allow_pickle=True)
        if score_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        green_points = np.load(dataset_name + "/st" + str(student) + "/points/" + str(c) + "_green.npy", allow_pickle=True)
        if green_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        green_points = green_points[np.argmax(score_points)]
        red_points = np.load(dataset_name + "/st" + str(student) + "/points/" + str(c) + "_red.npy", allow_pickle=True)
        if red_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        red_points = red_points[np.argmax(score_points)]
        msk = []  # masks for each samples
        gp = []  # green points
        rp = []  # red points
        
        ## Load Image
        image = names[c]  # samples c
        if len(image.shape) == 2:
                image = cv2.cvtColor((np.array(((image + 1) / 2) * 255, dtype='uint8')), cv2.COLOR_GRAY2RGB)
        imshape = image.shape[0],image.shape[1]
        if np.max(image) < 2:
            image = np.array(((image + 1) / 2) * 255, dtype='uint8')    
            
        ## Downsample the image
        imageBlur = cv2.GaussianBlur(image,(blur_ratio,blur_ratio),0)
        
        ## Load Label
        label = labels[c]  # GT for sample c
        label = label == 1
        mask = 0
        
        ## Set Downsampled Image
        time_part1i = time.time()
        
        predictor.set_image(imageBlur)

        time_part1 = time.time() - time_part1i
        
        ## Concatenate Green and Red Prompts
        x,y = 0,0
        green = []
        greenx = []
        greeny = []
        for g in green_points:
            x = g[0]
            y = g[1]
            green.append((x, y))
            greenx.append(x)
            greeny.append(y)
        
        red = []
        redx = []
        redy = []
        for r in red_points:
            x = r[0]
            y = r[1]
            red.append((x, y))
            redx.append(x)
            redy.append(y)
    
        score = 0.
        time_part2 = 0.
        if green and red:
            input_point = np.concatenate((green, red))
            input_label = np.concatenate(([1] * len(green), [0] * len(red)))

            time_part2i = time.time()
            masks, scores, logits = predictor.predict(
                point_coords=input_point,
                point_labels=input_label,
                multimask_output=True,
            )
            time_part2 = time.time() - time_part2i
            
            ## Score calculation
            mask = masks[0]
            intersection = (mask & label).sum()
            union = (mask | label).sum()
            if intersection == 0:
                score = 0
            else:
                score = intersection / union

        ws['A' + str(c + 2)] = c+1
        ws['B' + str(c + 2)] = score
        ws['C' + str(c + 2)] = time_part1
        ws['D' + str(c + 2)] = time_part2
        c += 1
        
        fig, ax = plt.subplots(1, 3, figsize=(15, 7))
        if green and red:
            ax[0].imshow(imageBlur)
            ax[1].imshow(label)
            ax[2].imshow(image)
            show_mask(mask, ax[1])
            show_mask(mask, ax[2])
            ax[0].plot(greenx, greeny, 'go', markersize=5)
            ax[0].plot(redx, redy, 'ro', markersize=5)
            plt.draw()
        
    wb.save(os.path.join(down_name, 'st' + str(student) + down_name + '.xlsx'))



