# -*- coding: utf-8 -*-
"""
Created on Mon Feb 19 14:25:18 2024

@author: Mohammed
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 01:47:37 2024

@author: Mohammed
"""

import sys
from time import sleep

using_colab = False
import numpy as np
import torch
import matplotlib.pyplot as plt
import cv2
from openpyxl import Workbook
import statistics
import os
import random
from IPython import get_ipython
from matplotlib.backend_bases import MouseButton
import time
from PIL import Image
from segment_anything import sam_model_registry, SamPredictor
import matplotlib

plt.rcParams['keymap.grid'].remove('g')
plt.rcParams['keymap.home'].remove('r')

MEDIUM_STAR_SIZE = 50 
MEDIUM_GREEN_RED_DOT_SIZE = 5
SMALL_STAR_SIZE = 10
SMALL_GREEN_RED_DOT_SIZE = 2

MEDIUM_DOT_SIZE_MODE = False
SMALL_DOT_SIZE_MODE = True
dot_size_toggle = SMALL_DOT_SIZE_MODE # small dot size by default
GREEN_COLOR = '#00f700'
RED_COLOR = '#ff1919'


sys.path.append("..")

try:
    matplotlib.use('Qt5Agg')
except:
    matplotlib.use('TkAgg')

sam_checkpoint = 'sam_vit_h_4b8939.pth'
model_type = "vit_h"

device = "cuda" if torch.cuda.is_available() else 'cpu'

sam = sam_model_registry[model_type](checkpoint=sam_checkpoint)
sam.to(device=device)

predictor = SamPredictor(sam)

names = np.load("samples.npy", allow_pickle=True)
labels = np.load("labels.npy", allow_pickle=True)

dataset_name = input("What dataset are you running?")
name = dataset_name + "_Timed_DownSampled"
if not os.path.exists(name):
    os.makedirs(name)


#%%

#### change that later

c=0

f = False
compressions = input("How many times do you want to compress the image? ")
## start looping through samples:
for student in np.arange(1,2):
    print("Student: ", student) 
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Sample Number'
    ws['B1'] = 'Scores'
    ws['C1'] = 'Times'
    c=0
    num_files = 0
    dir_path = dataset_name + "/st" + str(student) + "/scores/"
    # Iterate directory
    for path in os.listdir(dir_path):
        # check if current path is a file
        if os.path.isfile(os.path.join(dir_path, path)):
            num_files += 1
    while c < 5 and not f:
        print("Sample:", c)
        score_points = np.load("Cat/st" + str(student) + "/scores/" + str(c) + "score.npy", allow_pickle=True)
        if score_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        green_points = np.load("Cat/st" + str(student) + "/points/" + str(c) + "_green.npy", allow_pickle=True)
        if green_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        green_points = green_points[np.argmax(score_points)]
        red_points = np.load("Cat/st" + str(student) + "/points/" + str(c) + "_red.npy", allow_pickle=True)
        if red_points.size == 0:
            ws['A' + str(c + 2)] = str(c)
            ws['B' + str(c + 2)] = str(0)  # samples name on excel
            ws['C' + str(c + 2)] = str(0)
            c += 1
            continue
        red_points = red_points[np.argmax(score_points)]
        msk = []  # masks for each samples

        gp = []  # green points
        rp = []  # red points
        image = names[c]  # samples c
        if len(image.shape) == 2:
            image = cv2.cvtColor((np.array(((image + 1) / 2) * 255, dtype='uint8')), cv2.COLOR_GRAY2RGB)
        imshape = image.shape[0],image.shape[1]
        #maxPool = torch.nn.MaxPool2d(2,2,return_indices=True)
        maxPool = torch.nn.AvgPool2d(2,2)
        imageTensor = torch.from_numpy(image)
        imageTensor = imageTensor.permute(2, 0, 1)
        imageTensor = imageTensor.float().unsqueeze(dim=0)
        #imageTensorDone,indices = maxPool(imageTensor)
        for compr in range(int(compressions)):
            imageTensor = maxPool(imageTensor)

        no_batch = imageTensor.byte().squeeze(dim=0)

        # Unpermute
        imTens = no_batch.permute(1, 2, 0)
        imageDown = imTens.numpy()
        
        label = labels[c]  # GT for sample c
        rmv = False
        mask = 0
        # image=np.array(((image+1)/2)*255,dtype='uint8')
        time_part1i = time.time()
        
        predictor.set_image(imageDown)

        time_part1 = time.time() - time_part1i
        inc = ""
        co = 0
        bs = 0
        score = []
        round=[0,0]
        stdx = []
        stdy = []
        ng = []
        nr = []
        green = []
        red = []
        greenx = []
        times = []
        
        redx = []
        greeny = []
        redy = []
        
        # label=plt.imread('C:/Users/Mohammed/Downloads/labels/'+labels[c])i9i
        label = label == 1

        # matplotlib.use('TkAgg')

        while inc != "y":
            s = 0  # this is for the score
            count = 1  # to count the score max
            lessfive = 0
            current_color = 'green'
            dot_size_toggle = SMALL_DOT_SIZE_MODE # default will be small dot, not medium
            current_star_size = SMALL_STAR_SIZE
            current_green_red_dot_size = SMALL_GREEN_RED_DOT_SIZE
            # get_ipython().run_line_magic('matplotlib', 'qt')

            def addPointsPreventDupes():
                x,y = 0,0
                
                seen = set(green)
                for g in green_points:
                    x = g[0] / 2**int(compressions)
                    y = g[1] / 2**int(compressions)
                    
                    if ((x,y)) not in green:
                        seen.add((x,y))
                        green.append((x, y))
                        greenx.append(x)
                        greeny.append(y)
                
                seen = set(red)
                for r in red_points:
                    x = r[0] / 2**int(compressions)
                    y = r[1] / 2**int(compressions)
                    if ((x,y)) not in seen:
                        seen.add((x,y))
                        red.append((x, y))
                        redx.append(x)
                        redy.append(y)
            
                if green and red:
                    global s
                    #print("green:", green)
                    #print("red:", red)

                    input_point = np.concatenate((green, red))
                    input_label = np.concatenate(([1] * len(green), [0] * len(red)))

                    time_part2i = time.time()
                    
                    
                    # Get the time in seconds 
                    # since the epoch 
                    
                    masks, scores, logits = predictor.predict(
                        point_coords=input_point,
                        point_labels=input_label,
                        multimask_output=True,
                    )
                    time_part2 = time.time() - time_part2i
                    times.append(time_part1 + time_part2)
                    
                    # Print the time  
                # print("Time in seconds since the epoch:", time_sec) 
                    mask = masks[0]
                    maxunPool = torch.nn.Upsample(size=imshape, mode='nearest')
                    imageTensor = torch.from_numpy(mask)
                    imageTensor = imageTensor.float().unsqueeze(dim=0)
                    imageTensor = imageTensor.float().unsqueeze(dim=0)
                    imageTensorDone = maxunPool(imageTensor)
                    no_batch = imageTensorDone.long().squeeze(dim=0)
                    imTens = no_batch.long().squeeze(dim=0)
                    imageUp = imTens.numpy()
                    mask = imageUp.astype(int)

                    # get_ipython().run_line_magic('matplotlib', 'inline')
                    
                    intersection = (mask & label).sum()
                    union = (mask | label).sum()
                    if intersection == 0:
                        s = 0
                    else:
                        s = intersection / union
                    # ws[chr(68)+str(c+2)]=str(bs) # start at cell D(c)
                    #show_points(input_point, input_label, ax[2], marker_size = current_star_size)
                    msg = ""

                    if len(score[round[0]:]) == 0:
                        maxx = 0
                    else:
                        maxx = max(score[round[0]:])
                        print("maxx",maxx)
                    score.append(s)
                    gp.append(np.multiply(green, 1))

                    rp.append(np.multiply(red, 1))
                    ng.append(len(greenx))
                    nr.append(len(redx))
                    grx = np.concatenate([greenx, redx])
                    gry = np.concatenate([greeny, redy])

                    stdx.append(statistics.pstdev(grx.astype(int).tolist()))
                    stdy.append(statistics.pstdev(gry.astype(int).tolist()))
                
                    plt.title(f"Score: {(intersection / union):.3f}" + msg, fontsize=13)
                    ## saving masks, scores, points and other stats: 
                    msk.append(np.multiply(mask, 5))
                    #print("less than best score", lessfive)
                    #print("scores:", score[0])
                    #print("times:", times[0] )

            def addPoints():
                x,y = 0,0
                for g in green_points:
                    x = g[0] / 2**int(compressions)
                    y = g[1] / 2**int(compressions)
                    green.append((x, y))
                    greenx.append(x)
                    greeny.append(y)
                
                for r in red_points:
                    x = r[0] / 2**int(compressions)
                    y = r[1] / 2**int(compressions)
                    red.append((x, y))
                    redx.append(x)
                    redy.append(y)
            
                if green and red:
                    global s

                    input_point = np.concatenate((green, red))
                    input_label = np.concatenate(([1] * len(green), [0] * len(red)))

                    time_part2i = time.time()

                    
                    masks, scores, logits = predictor.predict(
                        point_coords=input_point,
                        point_labels=input_label,
                        multimask_output=True,
                    )
                    time_part2 = time.time() - time_part2i
                    times.append(time_part1 + time_part2)
                    

                    mask = masks[0]
                    maxunPool = torch.nn.Upsample(size=imshape, mode='nearest')
                    imageTensor = torch.from_numpy(mask)
                    imageTensor = imageTensor.float().unsqueeze(dim=0)
                    imageTensor = imageTensor.float().unsqueeze(dim=0)
                    imageTensorDone = maxunPool(imageTensor)
                    no_batch = imageTensorDone.long().squeeze(dim=0)
                    imTens = no_batch.long().squeeze(dim=0)
                    imageUp = imTens.numpy()
                    mask = imageUp.astype(int)
                    
                    intersection = (mask & label).sum()
                    union = (mask | label).sum()
                    if intersection == 0:
                        s = 0
                    else:
                        s = intersection / union
                    msg = ""

                    if len(score[round[0]:]) == 0:
                        maxx = 0
                    else:
                        maxx = max(score[round[0]:])
                        print("maxx",maxx)
                    score.append(s)
                    gp.append(np.multiply(green, 1))

                    rp.append(np.multiply(red, 1))
                    ng.append(len(greenx))
                    nr.append(len(redx))
                    grx = np.concatenate([greenx, redx])
                    gry = np.concatenate([greeny, redy])

                    stdx.append(statistics.pstdev(grx.astype(int).tolist()))
                    stdy.append(statistics.pstdev(gry.astype(int).tolist()))
                
                    plt.title(f"Score: {(intersection / union):.3f}" + msg, fontsize=13)
                    ## saving masks, scores, points and other stats: 
                    msk.append(np.multiply(mask, 5))

            addPoints()
            inc = "y"

        indx = np.argsort(-np.array(score))
        sscore = np.array(score)[indx]
        times_2 = np.array(times)[indx]
        snr = np.array(nr)[indx]
        sstdx = np.array(stdx)[indx]
        sstdy = np.array(stdy)[indx]
        ws['A' + str(c + 2)] = str(c)
        ws['B' + str(c + 2)] = sscore[0]  # samples name on excel
        ws['C' + str(c + 2)] = times_2[0]
        c += 1
        
    wb.save(os.path.join(name, 'st' + str(student) + name + '.xlsx'))



